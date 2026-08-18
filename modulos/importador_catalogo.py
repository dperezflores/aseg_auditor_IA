from __future__ import annotations

import hashlib
import json
import re
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


HOJAS = {
    "CATALOGO_DOCUMENTOS": "clave_catalogo",
    "FORMATOS_DOCUMENTO": "formato_id",
    "PROCEDIMIENTOS": "procedimiento_id",
    "CAMPOS_EXTRACCION": "campo_id",
}

REQUERIDOS = {
    "CATALOGO_DOCUMENTOS": {
        "clave_catalogo", "orden_en_procedimiento", "activo", "version_catalogo",
        "procedimiento", "etapa", "tipo_documental", "codigo_base",
        "nombre_documento", "obligatoriedad", "condicion_aplicabilidad",
        "admite_multiples", "riesgo_predeterminado", "regla_formatos",
        "criterios_identificacion_ia", "datos_clave_a_validar",
        "fundamento_normativo", "observaciones", "estado_revision", "fuente_origen",
    },
    "FORMATOS_DOCUMENTO": {
        "formato_id", "clave_catalogo", "extension", "mime_type",
        "nombre_por_formato", "modalidad_formato", "activo", "estado_revision",
        "observaciones", "fuente_origen",
    },
    "PROCEDIMIENTOS": {
        "procedimiento_id", "tipo_documental", "orden", "procedimiento_validacion",
        "riesgo_codigo", "activo", "estado_revision", "observaciones",
    },
    "CAMPOS_EXTRACCION": {
        "campo_id", "tipo_documental", "orden_salida", "nombre_tecnico",
        "etiqueta_salida", "tipo_dato", "obligatorio_ia",
        "instruccion_extraccion", "mostrar_en_detalle", "estado_revision",
        "observaciones",
    },
}


@dataclass(frozen=True)
class ResultadoValidacion:
    datos: dict[str, list[dict[str, Any]]]
    resumen: dict[str, int]
    advertencias: list[str]
    errores: list[str]
    archivo_sha256: str

    @property
    def valido(self) -> bool:
        return not self.errores


def _texto(valor: Any) -> str:
    if valor is None:
        return ""
    if isinstance(valor, float) and valor.is_integer():
        return str(int(valor))
    return str(valor).strip()


def _booleano(valor: Any, etiqueta: str) -> bool:
    normalizado = _texto(valor).casefold()
    if normalizado in {"sí", "si", "true", "1"}:
        return True
    if normalizado in {"no", "false", "0"}:
        return False
    raise ValueError(f"{etiqueta}: se esperaba Sí o No y se recibió {valor!r}")


def _entero(valor: Any, etiqueta: str) -> int:
    try:
        numero = int(valor)
    except (TypeError, ValueError) as exc:
        raise ValueError(f"{etiqueta}: se esperaba un número entero") from exc
    if numero <= 0:
        raise ValueError(f"{etiqueta}: debe ser mayor que cero")
    return numero


def _fecha(valor: Any, etiqueta: str) -> date | None:
    if valor in (None, ""):
        return None
    if isinstance(valor, datetime):
        return valor.date()
    if isinstance(valor, date):
        return valor
    texto = _texto(valor)
    try:
        return date.fromisoformat(texto[:10])
    except ValueError as exc:
        raise ValueError(f"{etiqueta}: use una fecha válida con formato AAAA-MM-DD") from exc


def _leer_hoja(hoja, nombre: str, clave: str) -> list[dict[str, Any]]:
    encabezados: list[str] | None = None
    fila_encabezados = 0
    for numero, fila in enumerate(hoja.iter_rows(values_only=True), start=1):
        valores = [_texto(valor) for valor in fila]
        if clave in valores:
            encabezados = valores
            fila_encabezados = numero
            break
    if not encabezados:
        raise ValueError(f"{nombre}: no se encontró la columna principal {clave!r}")

    faltantes = REQUERIDOS[nombre] - set(encabezados)
    if faltantes:
        raise ValueError(f"{nombre}: faltan columnas: {', '.join(sorted(faltantes))}")

    registros: list[dict[str, Any]] = []
    for numero, fila in enumerate(
        hoja.iter_rows(min_row=fila_encabezados + 1, values_only=True),
        start=fila_encabezados + 1,
    ):
        valores = list(fila)
        registro = {
            encabezado: valores[indice] if indice < len(valores) else None
            for indice, encabezado in enumerate(encabezados)
            if encabezado
        }
        if not _texto(registro.get(clave)):
            continue
        registro["_fila"] = numero
        registros.append(registro)
    return registros


def _duplicados(registros: list[dict[str, Any]], clave: str) -> list[str]:
    vistos: set[str] = set()
    repetidos: set[str] = set()
    for registro in registros:
        valor = _texto(registro.get(clave))
        if valor in vistos:
            repetidos.add(valor)
        vistos.add(valor)
    return sorted(repetidos)


def leer_y_validar_catalogo(ruta: str | Path) -> ResultadoValidacion:
    ruta = Path(ruta)
    archivo_sha256 = hashlib.sha256(ruta.read_bytes()).hexdigest()
    libro = load_workbook(ruta, read_only=True, data_only=True)
    errores: list[str] = []
    advertencias: list[str] = []
    datos: dict[str, list[dict[str, Any]]] = {}

    try:
        for nombre, clave in HOJAS.items():
            if nombre not in libro.sheetnames:
                errores.append(f"Falta la hoja obligatoria {nombre}")
                datos[nombre] = []
                continue
            try:
                datos[nombre] = _leer_hoja(libro[nombre], nombre, clave)
            except ValueError as exc:
                errores.append(str(exc))
                datos[nombre] = []
    finally:
        libro.close()

    for nombre, clave in HOJAS.items():
        for duplicado in _duplicados(datos[nombre], clave):
            errores.append(f"{nombre}: {clave} duplicado: {duplicado}")

    catalogo = datos["CATALOGO_DOCUMENTOS"]
    claves = {_texto(fila["clave_catalogo"]) for fila in catalogo}
    tipos_etapas: dict[str, str] = {}
    for fila in catalogo:
        numero = fila["_fila"]
        prefijo = f"CATALOGO_DOCUMENTOS fila {numero}"
        try:
            fila["clave_catalogo"] = _texto(fila["clave_catalogo"])
            fila["orden_en_procedimiento"] = _entero(fila["orden_en_procedimiento"], prefijo)
            fila["activo"] = _booleano(fila["activo"], prefijo)
            fila["version_catalogo"] = _texto(fila["version_catalogo"])
            fila["procedimiento"] = _texto(fila["procedimiento"]).upper()
            fila["etapa"] = _texto(fila["etapa"]).upper()
            fila["tipo_documental"] = _texto(fila["tipo_documental"]).upper()
            fila["codigo_base"] = _texto(fila["codigo_base"]).upper()
            fila["nombre_documento"] = _texto(fila["nombre_documento"])
            fila["obligatoriedad"] = _texto(fila["obligatoriedad"])
            fila["condicion_aplicabilidad"] = _texto(fila["condicion_aplicabilidad"])
            fila["admite_multiples"] = _booleano(fila["admite_multiples"], prefijo)
            fila["patron_consecutivo"] = _texto(fila.get("patron_consecutivo")) or None
            fila["riesgo_predeterminado"] = _texto(fila["riesgo_predeterminado"])
            fila["regla_formatos"] = _texto(fila["regla_formatos"])
            fila["criterios_identificacion_ia"] = _texto(fila["criterios_identificacion_ia"])
            fila["datos_clave_a_validar"] = _texto(fila["datos_clave_a_validar"])
            fila["fundamento_normativo"] = _texto(fila["fundamento_normativo"])
            fila["vigencia_desde"] = _fecha(fila.get("vigencia_desde"), prefijo)
            fila["vigencia_hasta"] = _fecha(fila.get("vigencia_hasta"), prefijo)
            fila["observaciones"] = _texto(fila["observaciones"])
            fila["estado_revision"] = _texto(fila["estado_revision"])
            fila["fuente_origen"] = _texto(fila["fuente_origen"])
        except ValueError as exc:
            errores.append(str(exc))
            continue

        if fila["procedimiento"] not in {"LPU", "LSI", "DIR", "ADM"}:
            errores.append(f"{prefijo}: procedimiento no permitido")
        if fila["etapa"] not in {"PPP", "ADJ", "CNT", "EJE", "ETR"}:
            errores.append(f"{prefijo}: etapa no permitida")
        if fila["estado_revision"] not in {"Pendiente", "Revisado", "Aprobado"}:
            errores.append(f"{prefijo}: estado_revision no permitido")
        if not fila["admite_multiples"] and fila["patron_consecutivo"]:
            errores.append(f"{prefijo}: tiene patrón consecutivo, pero no admite múltiples")
        if fila["admite_multiples"] and not fila["patron_consecutivo"]:
            errores.append(f"{prefijo}: admite múltiples y requiere patrón consecutivo")
        if fila["vigencia_desde"] and fila["vigencia_hasta"] and fila["vigencia_hasta"] < fila["vigencia_desde"]:
            errores.append(f"{prefijo}: vigencia_hasta es anterior a vigencia_desde")
        if fila["patron_consecutivo"] and re.search(r"\.[A-Za-z0-9]+$", fila["patron_consecutivo"]):
            advertencias.append(f"{fila['clave_catalogo']}: el patrón consecutivo incluye extensión")
        if re.search(r"_\d+$", fila["clave_catalogo"]):
            advertencias.append(f"{fila['clave_catalogo']}: la clave del catálogo termina en consecutivo")
        if fila["activo"] and fila["estado_revision"] != "Aprobado":
            advertencias.append(f"{fila['clave_catalogo']}: está activo, pero aún no está aprobado")

        tipo = fila["tipo_documental"]
        etapa_anterior = tipos_etapas.setdefault(tipo, fila["etapa"])
        if etapa_anterior != fila["etapa"]:
            errores.append(f"{tipo}: aparece en más de una etapa")

    for fila in datos["FORMATOS_DOCUMENTO"]:
        prefijo = f"FORMATOS_DOCUMENTO fila {fila['_fila']}"
        try:
            fila["formato_id"] = _texto(fila["formato_id"])
            fila["clave_catalogo"] = _texto(fila["clave_catalogo"])
            fila["extension"] = _texto(fila["extension"]).lower().lstrip(".")
            fila["mime_type"] = _texto(fila["mime_type"])
            fila["nombre_por_formato"] = _texto(fila["nombre_por_formato"])
            fila["modalidad_formato"] = _texto(fila["modalidad_formato"])
            fila["activo"] = _booleano(fila["activo"], prefijo)
            fila["estado_revision"] = _texto(fila["estado_revision"])
            fila["observaciones"] = _texto(fila["observaciones"])
            fila["fuente_origen"] = _texto(fila["fuente_origen"])
        except ValueError as exc:
            errores.append(str(exc))
            continue
        if fila["clave_catalogo"] not in claves:
            errores.append(f"{prefijo}: clave_catalogo inexistente: {fila['clave_catalogo']}")

    tipos = set(tipos_etapas)
    for fila in datos["PROCEDIMIENTOS"]:
        prefijo = f"PROCEDIMIENTOS fila {fila['_fila']}"
        try:
            fila["procedimiento_id"] = _texto(fila["procedimiento_id"])
            fila["tipo_documental"] = _texto(fila["tipo_documental"]).upper()
            fila["orden"] = _entero(fila["orden"], prefijo)
            fila["procedimiento_validacion"] = _texto(fila["procedimiento_validacion"])
            fila["resultado_esperado"] = _texto(fila.get("resultado_esperado")) or None
            fila["evidencia_requerida"] = _texto(fila.get("evidencia_requerida")) or None
            fila["riesgo_codigo"] = _texto(fila["riesgo_codigo"])
            fila["activo"] = _booleano(fila["activo"], prefijo)
            fila["estado_revision"] = _texto(fila["estado_revision"])
            fila["observaciones"] = _texto(fila["observaciones"])
        except ValueError as exc:
            errores.append(str(exc))
            continue
        if fila["tipo_documental"] not in tipos:
            errores.append(f"{prefijo}: tipo_documental inexistente: {fila['tipo_documental']}")

    for fila in datos["CAMPOS_EXTRACCION"]:
        prefijo = f"CAMPOS_EXTRACCION fila {fila['_fila']}"
        try:
            fila["campo_id"] = _texto(fila["campo_id"])
            fila["tipo_documental"] = _texto(fila["tipo_documental"]).upper()
            fila["orden_salida"] = _entero(fila["orden_salida"], prefijo)
            fila["nombre_tecnico"] = _texto(fila["nombre_tecnico"])
            fila["etiqueta_salida"] = _texto(fila["etiqueta_salida"])
            fila["tipo_dato"] = _texto(fila["tipo_dato"])
            fila["obligatorio_ia"] = _booleano(fila["obligatorio_ia"], prefijo)
            fila["instruccion_extraccion"] = _texto(fila["instruccion_extraccion"])
            fila["mostrar_en_detalle"] = _booleano(fila["mostrar_en_detalle"], prefijo)
            fila["celda_pt"] = _texto(fila.get("celda_pt")) or None
            fila["estado_revision"] = _texto(fila["estado_revision"])
            fila["observaciones"] = _texto(fila["observaciones"])
        except ValueError as exc:
            errores.append(str(exc))
            continue
        if fila["tipo_documental"] not in tipos:
            errores.append(f"{prefijo}: tipo_documental inexistente: {fila['tipo_documental']}")

    resumen = {
        "documentos": len(catalogo),
        "tipos_documentales": len(tipos),
        "formatos": len(datos["FORMATOS_DOCUMENTO"]),
        "procedimientos": len(datos["PROCEDIMIENTOS"]),
        "campos_extraccion": len(datos["CAMPOS_EXTRACCION"]),
        "aprobados": sum(fila.get("estado_revision") == "Aprobado" for fila in catalogo),
    }
    return ResultadoValidacion(datos, resumen, advertencias, errores, archivo_sha256)


def importar_catalogo(
    resultado: ResultadoValidacion,
    archivo_nombre: str,
    database_url: str,
    activar: bool = False,
) -> tuple[str, bool]:
    if not resultado.valido:
        raise ValueError("El catálogo contiene errores y no puede importarse")

    import psycopg
    from psycopg.types.json import Jsonb

    with psycopg.connect(database_url, connect_timeout=15) as conn:
        with conn.cursor() as cur:
            cur.execute(
                "SELECT id, estado FROM catalogo_importaciones WHERE archivo_sha256 = %s",
                (resultado.archivo_sha256,),
            )
            existente = cur.fetchone()
            if existente and existente[1] == "ACTIVO":
                return str(existente[0]), True

            if existente:
                importacion_id = existente[0]
                cur.execute("DELETE FROM campos_extraccion WHERE importacion_id = %s", (importacion_id,))
                cur.execute("DELETE FROM procedimientos_validacion WHERE importacion_id = %s", (importacion_id,))
                cur.execute("DELETE FROM catalogo_documentos WHERE importacion_id = %s", (importacion_id,))
                cur.execute("DELETE FROM tipos_documentales WHERE importacion_id = %s", (importacion_id,))
                cur.execute(
                    """
                    UPDATE catalogo_importaciones
                    SET archivo_nombre = %s, estado = 'VALIDADO', resumen = %s,
                        advertencias = %s, activado_en = NULL
                    WHERE id = %s
                    """,
                    (archivo_nombre, Jsonb(resultado.resumen), Jsonb(resultado.advertencias), importacion_id),
                )
            else:
                cur.execute(
                    """
                    INSERT INTO catalogo_importaciones (
                        archivo_nombre, archivo_sha256, estado, resumen, advertencias
                    ) VALUES (%s, %s, 'VALIDADO', %s, %s)
                    RETURNING id
                    """,
                    (
                        archivo_nombre,
                        resultado.archivo_sha256,
                        Jsonb(resultado.resumen),
                        Jsonb(resultado.advertencias),
                    ),
                )
                importacion_id = cur.fetchone()[0]

            catalogo = resultado.datos["CATALOGO_DOCUMENTOS"]
            tipos_etapas = {
                fila["tipo_documental"]: fila["etapa"] for fila in catalogo
            }
            cur.executemany(
                "INSERT INTO tipos_documentales (importacion_id, tipo_documental, etapa) VALUES (%s, %s, %s)",
                [(importacion_id, tipo, etapa) for tipo, etapa in tipos_etapas.items()],
            )

            ids_catalogo: dict[str, Any] = {}
            for fila in catalogo:
                cur.execute(
                    """
                    INSERT INTO catalogo_documentos (
                        importacion_id, clave_catalogo, orden_en_procedimiento, activo,
                        version_catalogo, procedimiento, etapa, tipo_documental, codigo_base,
                        nombre_documento, obligatoriedad, condicion_aplicabilidad,
                        admite_multiples, patron_consecutivo, riesgo_predeterminado,
                        regla_formatos, criterios_identificacion_ia, datos_clave_a_validar,
                        fundamento_normativo, vigencia_desde, vigencia_hasta, observaciones,
                        estado_revision, fuente_origen
                    ) VALUES (
                        %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s,
                        %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s
                    ) RETURNING id
                    """,
                    (
                        importacion_id, fila["clave_catalogo"], fila["orden_en_procedimiento"],
                        fila["activo"], fila["version_catalogo"], fila["procedimiento"],
                        fila["etapa"], fila["tipo_documental"], fila["codigo_base"],
                        fila["nombre_documento"], fila["obligatoriedad"],
                        fila["condicion_aplicabilidad"], fila["admite_multiples"],
                        fila["patron_consecutivo"], fila["riesgo_predeterminado"],
                        fila["regla_formatos"], fila["criterios_identificacion_ia"],
                        fila["datos_clave_a_validar"], fila["fundamento_normativo"],
                        fila["vigencia_desde"], fila["vigencia_hasta"], fila["observaciones"],
                        fila["estado_revision"], fila["fuente_origen"],
                    ),
                )
                ids_catalogo[fila["clave_catalogo"]] = cur.fetchone()[0]

            cur.executemany(
                """
                INSERT INTO formatos_documento (
                    importacion_id, formato_id, catalogo_documento_id, extension, mime_type,
                    nombre_por_formato, modalidad_formato, activo, estado_revision,
                    observaciones, fuente_origen
                ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                """,
                [
                    (
                        importacion_id, fila["formato_id"], ids_catalogo[fila["clave_catalogo"]],
                        fila["extension"], fila["mime_type"], fila["nombre_por_formato"],
                        fila["modalidad_formato"], fila["activo"], fila["estado_revision"],
                        fila["observaciones"], fila["fuente_origen"],
                    )
                    for fila in resultado.datos["FORMATOS_DOCUMENTO"]
                ],
            )

            cur.executemany(
                """
                INSERT INTO procedimientos_validacion (
                    importacion_id, procedimiento_id, tipo_documental, orden,
                    procedimiento_validacion, resultado_esperado, evidencia_requerida,
                    riesgo_codigo, activo, estado_revision, observaciones
                ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                """,
                [
                    (
                        importacion_id, fila["procedimiento_id"], fila["tipo_documental"],
                        fila["orden"], fila["procedimiento_validacion"],
                        fila["resultado_esperado"], fila["evidencia_requerida"],
                        fila["riesgo_codigo"], fila["activo"], fila["estado_revision"],
                        fila["observaciones"],
                    )
                    for fila in resultado.datos["PROCEDIMIENTOS"]
                ],
            )

            cur.executemany(
                """
                INSERT INTO campos_extraccion (
                    importacion_id, campo_id, tipo_documental, orden_salida,
                    nombre_tecnico, etiqueta_salida, tipo_dato, obligatorio_ia,
                    instruccion_extraccion, mostrar_en_detalle, celda_pt,
                    estado_revision, observaciones
                ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                """,
                [
                    (
                        importacion_id, fila["campo_id"], fila["tipo_documental"],
                        fila["orden_salida"], fila["nombre_tecnico"], fila["etiqueta_salida"],
                        fila["tipo_dato"], fila["obligatorio_ia"],
                        fila["instruccion_extraccion"], fila["mostrar_en_detalle"],
                        fila["celda_pt"], fila["estado_revision"], fila["observaciones"],
                    )
                    for fila in resultado.datos["CAMPOS_EXTRACCION"]
                ],
            )

            if activar:
                cur.execute(
                    "UPDATE catalogo_importaciones SET estado = 'ARCHIVADO' WHERE estado = 'ACTIVO' AND id <> %s",
                    (importacion_id,),
                )
                cur.execute(
                    "UPDATE catalogo_importaciones SET estado = 'ACTIVO', activado_en = now() WHERE id = %s",
                    (importacion_id,),
                )
        conn.commit()
    return str(importacion_id), False


def resumen_json(resultado: ResultadoValidacion) -> str:
    return json.dumps(
        {
            "valido": resultado.valido,
            "resumen": resultado.resumen,
            "errores": resultado.errores,
            "advertencias": resultado.advertencias,
            "sha256": resultado.archivo_sha256,
        },
        ensure_ascii=False,
        indent=2,
        default=str,
    )

